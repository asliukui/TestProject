from openpyxl import load_workbook

urlExle=r"C:\Users\asliu\Desktop\aaa.xlsx"
# wb = Workbook()
# print(wb.active)
# wb = load_workbook(urlExle)
# wa = wb.active
# print(wb.sheetnames)


# wb.save(r"C:\Users\asliu\Desktop\bbb.xlsx") #保存


# 加载 Excel 文件
wb = load_workbook(urlExle)

# 获取活动工作表
wa = wb.active
rd = wa.iter_rows(1,2)
print(rd)
ws = wb.sheetnames
ce = wa.cell(3,2)
print("cell(3,2)的值:",ce)

print(wb["asset_repay_plan"].values)
wa.auto_filter.ref= "A1:B3"
# 获取单元格对象
